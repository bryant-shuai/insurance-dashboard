import pandas as pd
import os
from openpyxl import load_workbook
import csv
import json
import logging

logger = logging.getLogger(__name__)

class ExcelParser:
    def __init__(self):
        self.insurances = []
        self.companies = {}
        self.raw_html = ""
    
    def parse_file(self, file_path):
        """解析Excel或CSV文件"""
        try:
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.csv':
                return self._parse_csv(file_path)
            else:
                return self._parse_excel(file_path)
                
        except Exception as e:
            logger.error(f"文件解析失败: {str(e)}")
            raise e
    
    def _parse_csv(self, file_path):
        """解析CSV文件"""
        try:
            # 读取CSV文件
            df = pd.read_csv(file_path, encoding='utf-8')
            json_data = df.values.tolist()
            
            # 生成简单的HTML表格
            html_table = '<table border="0" cellpadding="0" cellspacing="0">'
            for row in json_data:
                html_table += '<tr>'
                for cell in row:
                    html_table += f'<td>{cell}</td>'
                html_table += '</tr>'
            html_table += '</table>'
            self.raw_html = html_table
            
            return self._process_data(json_data)
            
        except Exception as e:
            logger.error(f"CSV解析失败: {str(e)}")
            raise e
    
    def _parse_excel(self, file_path):
        """解析Excel文件"""
        try:
            # 使用pandas读取Excel
            df = pd.read_excel(file_path)
            json_data = df.values.tolist()
            
            # 生成HTML表格
            wb = load_workbook(file_path)
            ws = wb.active
            self.raw_html = self._worksheet_to_html(ws)
            
            return self._process_data(json_data)
            
        except Exception as e:
            logger.error(f"Excel解析失败: {str(e)}")
            raise e
    
    def _worksheet_to_html(self, worksheet):
        """将worksheet转换为HTML表格"""
        html = '<table border="0" cellpadding="0" cellspacing="0">'
        for row in worksheet.iter_rows():
            html += '<tr>'
            for cell in row:
                html += f'<td>{cell.value if cell.value else ""}</td>'
            html += '</tr>'
        html += '</table>'
        return html
    
    def _process_data(self, json_data):
        """处理解析后的数据，提取表头和指标"""
        header_idx = -1
        metric_idx = -1
        
        # 查找表头行和指标行
        for i in range(min(20, len(json_data))):
            if i >= len(json_data):
                continue
            row = json_data[i] if json_data[i] else []
            str_row = ' '.join(str(cell) for cell in row if cell)
            
            if '非车险' in str_row:
                header_idx = i
                logger.debug(f"找到表头行: {i}")
            
            if '本期累计' in str_row or '同比增长' in str_row:
                metric_idx = i
                logger.debug(f"找到指标行: {i}")
        
        if header_idx == -1 or metric_idx == -1:
            raise ValueError('表头识别失败')
        
        headers = json_data[header_idx]
        metrics = json_data[metric_idx]
        
        # 构建列映射
        col_map = {}
        insurances = []
        
        # 提取险种列表（从第2列开始）
        for i in range(2, len(headers)):
            header = headers[i] if i < len(headers) else None
            if header and header not in ['地区', '公司名称']:
                ins_name = str(header).strip()
                if ins_name not in insurances:
                    insurances.append(ins_name)
                    col_map[ins_name] = {}
        
        # 提取每个险种的列索引
        ins_index = 0
        for i in range(2, len(metrics)):
            if i >= len(metrics) or ins_index >= len(insurances):
                break
                
            metric = metrics[i]
            if metric and ins_index < len(insurances):
                ins_name = insurances[ins_index]
                metric_str = str(metric).strip()
                
                if '本期累计' in metric_str:
                    col_map[ins_name]['p'] = i
                elif '同比增长' in metric_str:
                    col_map[ins_name]['g'] = i
                    ins_index += 1
        
        logger.debug(f"列映射: {col_map}")
        logger.debug(f"险种列表: {insurances}")
        
        # 提取公司数据
        companies = {}
        for i in range(metric_idx + 1, len(json_data)):
            row = json_data[i] if i < len(json_data) else []
            if not row or len(row) < 2:
                continue
                
            region = str(row[0]) if row[0] else ''
            name = str(row[1]) if row[1] else ''
            
            if '合计' in region or '合计' in name or not name:
                continue
            
            full_name = f"{region}-{name}" if region else name
            companies[full_name] = {}
            
            for ins in insurances:
                map_info = col_map.get(ins, {})
                if 'p' in map_info:
                    try:
                        p = float(row[map_info['p']]) if map_info['p'] < len(row) and row[map_info['p']] else 0
                        g_str = str(row[map_info['g']]).replace('%', '') if 'g' in map_info and map_info['g'] < len(row) else '0'
                        g = float(g_str) if g_str.replace('.', '').replace('-', '').isdigit() else 0
                        companies[full_name][ins] = {'premium': p, 'growth': g}
                    except (ValueError, IndexError):
                        companies[full_name][ins] = {'premium': 0, 'growth': 0}
        
        # 将非车险排在前面
        insurances.sort(key=lambda x: 0 if x == '非车险' else 1)
        
        self.insurances = insurances
        self.companies = companies
        
        return {
            'insurances': insurances,
            'companies': companies,
            'raw_html': self.raw_html
        }